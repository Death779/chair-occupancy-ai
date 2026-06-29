import os
import sqlite3
import urllib.request
from fpdf import FPDF
from datetime import datetime
from flask import Flask, render_template, request, send_file, redirect, url_for, jsonify
from ultralytics import YOLO
import cv2
import openpyxl

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'static/uploads'
app.config['RESULT_FOLDER'] = 'static/results'

# Создаем папки при старте
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)
os.makedirs(app.config['RESULT_FOLDER'], exist_ok=True)

# Ленивая загрузка моделей для экономии памяти
models = {}
def get_model(size):
    valid_sizes = ['n', 's', 'm']
    if size not in valid_sizes:
        size = 'n'
    if size not in models:
        models[size] = YOLO(f'yolov8{size}.pt')
    return models[size]

def init_db():
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    c.execute('''
        CREATE TABLE IF NOT EXISTS history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            timestamp TEXT,
            orig_path TEXT,
            res_path TEXT,
            persons INTEGER,
            chairs INTEGER,
            occupancy REAL
        )
    ''')
    # Добавляем колонку confidence, если её нет (для обратной совместимости)
    try:
        c.execute('ALTER TABLE history ADD COLUMN confidence REAL')
    except sqlite3.OperationalError:
        pass # Колонка уже существует
    conn.commit()
    conn.close()

init_db()

@app.route('/', methods=['GET'])
def index():
    return render_template('index.html')

@app.route('/upload', methods=['POST'])
def upload():
    if 'file' not in request.files:
        return jsonify({'error': 'Файл не найден'}), 400
    
    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'Файл не выбран'}), 400
    
    # Получаем параметры из запроса, устанавливаем значения по умолчанию
    confidence = float(request.form.get('confidence', 0.25))
    model_size = request.form.get('model_size', 'n')
    
    if file:
        filename = file.filename
        orig_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(orig_path)
        
        # Загружаем выбранную модель и запускаем инференс с заданным порогом
        model = get_model(model_size)
        results = model(orig_path, conf=confidence)
        
        persons = 0
        chairs = 0
        
        # Подсчет объектов нужных классов
        for r in results:
            boxes = r.boxes
            for box in boxes:
                cls_id = int(box.cls[0])
                class_name = model.names[cls_id]
                if class_name == 'person':
                    persons += 1
                elif class_name == 'chair':
                    chairs += 1
        
        # Сохранение изображения с bounding boxes
        res_filename = 'res_' + filename
        res_path = os.path.join(app.config['RESULT_FOLDER'], res_filename)
        
        res_img = results[0].plot()
        cv2.imwrite(res_path, res_img)
        
        # Обработка деления на ноль и ограничение > 100%
        if chairs == 0:
            occupancy = 0.0
            status = "Стулья не обнаружены"
            occupied_seats = 0
        else:
            occupied_seats = min(persons, chairs)
            occupancy = (occupied_seats / chairs) * 100.0
            if persons > chairs:
                status = "Зал переполнен"
            else:
                status = "Норма"
            
        occupancy = round(occupancy, 2)
        
        # Сохранение в базу данных
        conn = sqlite3.connect('database.db')
        c = conn.cursor()
        c.execute('INSERT INTO history (timestamp, orig_path, res_path, persons, chairs, occupancy, confidence) VALUES (?, ?, ?, ?, ?, ?, ?)',
                  (datetime.now().strftime("%Y-%m-%d %H:%M:%S"), orig_path, res_path, persons, chairs, occupancy, confidence))
        record_id = c.lastrowid
        conn.commit()
        conn.close()
        
        # Возвращаем JSON для обновления интерфейса без перезагрузки
        return jsonify({
            'record_id': record_id,
            'people_count': persons,
            'chairs_count': chairs,
            'occupied_seats': occupied_seats,
            'occupancy': occupancy,
            'status': status,
            'res_path': res_path,
            'orig_path': orig_path
        })

@app.route('/history')
def history():
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    c.execute('SELECT * FROM history ORDER BY id DESC')
    rows = c.fetchall()
    conn.close()
    return render_template('history.html', rows=rows)

@app.route('/export')
def export():
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    c.execute('SELECT * FROM history')
    rows = c.fetchall()
    conn.close()
    
    # Генерация Excel отчета
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "History Report"
    ws.append(["ID", "Дата и Время", "Оригинал", "Результат", "Кол-во людей", "Кол-во стульев", "Заполненность (%)", "Порог уверенности"])
    
    for row in rows:
        row_list = list(row)
        # Дополняем N/A если запись старая и нет колонки confidence
        if len(row_list) < 8:
            row_list.append('N/A')
        ws.append(row_list)
        
    export_path = "report.xlsx"
    wb.save(export_path)
    
    return send_file(export_path, as_attachment=True)

@app.route('/export_pdf')
def export_pdf():
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    c.execute('SELECT * FROM history')
    rows = c.fetchall()
    conn.close()

    font_path = "Roboto-Regular.ttf"
    if not os.path.exists(font_path):
        req = urllib.request.Request(
            "https://cdnjs.cloudflare.com/ajax/libs/pdfmake/0.1.66/fonts/Roboto/Roboto-Regular.ttf",
            headers={'User-Agent': 'Mozilla/5.0'}
        )
        with urllib.request.urlopen(req) as response, open(font_path, 'wb') as out_file:
            out_file.write(response.read())

    pdf = FPDF()
    pdf.add_page()
    pdf.add_font("Roboto", "", font_path)
    pdf.set_font("Roboto", size=14)
    pdf.cell(200, 10, text="Отчет по истории обработки", new_x="LMARGIN", new_y="NEXT", align="C")
    
    pdf.set_font("Roboto", size=10)
    
    col_widths = [10, 45, 25, 25, 35, 35]
    headers = ["ID", "Дата", "Люди", "Стулья", "Заполненность", "Порог"]
    for i in range(len(headers)):
        pdf.cell(col_widths[i], 10, text=headers[i], border=1, align="C")
    pdf.ln()
    
    for row in rows:
        row_list = list(row)
        if len(row_list) < 8:
            row_list.append('N/A')
            
        pdf.cell(col_widths[0], 10, text=str(row_list[0]), border=1, align="C")
        pdf.cell(col_widths[1], 10, text=str(row_list[1]), border=1, align="C")
        pdf.cell(col_widths[2], 10, text=str(row_list[4]), border=1, align="C")
        pdf.cell(col_widths[3], 10, text=str(row_list[5]), border=1, align="C")
        pdf.cell(col_widths[4], 10, text=f"{row_list[6]}%", border=1, align="C")
        pdf.cell(col_widths[5], 10, text=str(row_list[7]), border=1, align="C")
        pdf.ln()

    export_path = "report.pdf"
    pdf.output(export_path)
    return send_file(export_path, as_attachment=True)

@app.route('/export_single_pdf/<int:record_id>')
def export_single_pdf(record_id):
    conn = sqlite3.connect('database.db')
    c = conn.cursor()
    c.execute('SELECT * FROM history WHERE id = ?', (record_id,))
    row = c.fetchone()
    conn.close()

    if not row:
        return "Record not found", 404

    font_path = "Roboto-Regular.ttf"
    if not os.path.exists(font_path):
        req = urllib.request.Request(
            "https://cdnjs.cloudflare.com/ajax/libs/pdfmake/0.1.66/fonts/Roboto/Roboto-Regular.ttf",
            headers={'User-Agent': 'Mozilla/5.0'}
        )
        with urllib.request.urlopen(req) as response, open(font_path, 'wb') as out_file:
            out_file.write(response.read())

    pdf = FPDF()
    pdf.add_page()
    pdf.add_font("Roboto", "", font_path)
    pdf.set_font("Roboto", size=16)
    pdf.cell(200, 10, text="Отчет по обработке изображения", new_x="LMARGIN", new_y="NEXT", align="C")
    pdf.ln(10)
    
    pdf.set_font("Roboto", size=12)
    pdf.cell(200, 10, text=f"ID записи: {row[0]}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(200, 10, text=f"Дата и время: {row[1]}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(200, 10, text=f"Найдено людей: {row[4]}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(200, 10, text=f"Найдено стульев: {row[5]}", new_x="LMARGIN", new_y="NEXT")
    
    occupied_seats = min(row[4], row[5]) if row[5] > 0 else 0
    pdf.cell(200, 10, text=f"Занято мест: {occupied_seats}", new_x="LMARGIN", new_y="NEXT")
    pdf.cell(200, 10, text=f"Заполненность зала: {row[6]}%", new_x="LMARGIN", new_y="NEXT")
    
    if row[5] == 0:
        status = "Стулья не обнаружены"
    elif row[4] > row[5]:
        status = "Зал переполнен"
    else:
        status = "Норма"
        
    pdf.cell(200, 10, text=f"Статус: {status}", new_x="LMARGIN", new_y="NEXT")
    
    confidence = row[7] if len(row) > 7 and row[7] is not None else "N/A"
    pdf.cell(200, 10, text=f"Порог уверенности: {confidence}", new_x="LMARGIN", new_y="NEXT")
    
    pdf.ln(10)
    
    res_path = row[3]
    if os.path.exists(res_path):
        pdf.cell(200, 10, text="Результат детекции (с bounding boxes):", new_x="LMARGIN", new_y="NEXT")
        pdf.image(res_path, w=150)
        
    export_path = f"report_single_{record_id}.pdf"
    pdf.output(export_path)
    return send_file(export_path, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True, port=5000)
